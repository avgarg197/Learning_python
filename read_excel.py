import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError(
        "openpyxl is required to run this script. Install it with: pip install openpyxl"
    )


def read_excel_file(file_path: Path, sheet_name: str | None = None) -> None:
    workbook = load_workbook(filename=file_path, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
            )
        sheets = [workbook[sheet_name]]
    else:
        sheets = [workbook[workbook.sheetnames[0]]]

    for sheet in sheets:
        print(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            print("\t".join([str(cell) if cell is not None else "" for cell in row]))
        print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Read and print the contents of an Excel file (.xlsx)."
    )
    parser.add_argument("file", type=Path, help="Path to the Excel file")
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name to read (defaults to first sheet)",
    )
    args = parser.parse_args()

    if not args.file.exists():
        raise FileNotFoundError(f"Excel file not found: {args.file}")

    read_excel_file(args.file, args.sheet)
